import array
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from multiprocessing import Pool
from datetime import datetime
import concurrent.futures
import doctest

"""
vacancies.csv
Аналитик
"""

class Report:
    """Класс для вывода статистики по профессии в виде pdf с таблицами из excel и графиками

    Attributes:
        wb (Workbook): Рабочая книга excel
        ws (Worksheet): Активная страница
        ws.title (str): Название первой страницы
        ws1 (str): Название второй страницы
        border (Border): Рамки колонок
        fig, ax (any): Разметка для графиков
        profession (str): Профессия
        env (Environment): главный компонент Jinja
        html (Template): html страница
        profession (str): профессия
    """
    def __init__(self, profession):
        """Инициализирует объект Report

        Args:
            profession (str): Профессия
        """

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Статистика по годам"
        self.ws1 = self.wb.create_sheet("Статистика по городам")
        self.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        self.fig, self.ax = plt.subplots(nrows=2, ncols=2, figsize=(10, 10))
        self.profession = profession
        self.env = Environment(loader=FileSystemLoader('.'))
        self.html = self.env.get_template("main.html")

    def set_column_width(self, name_column, sheet, count_rows, start):
        """Выставляет каждому столбцу минимально возможную длину в зависимости самого большого по длине элемента

        Args:
            name_column (List): Список столбцов с названиями
            sheet (Worksheet): Активная страница
            count_rows (int): Количество строк
            start (str): Номер первого столбца в excel
        """

        for i in range(len(name_column)):
            max_len = 0
            for j in range(1, count_rows + 1):
                if max_len < len(str(sheet[chr(i + ord(start)) + str(j)].value)):
                    max_len = len(str(sheet[chr(i + ord(start)) + str(j)].value)) + 2
            sheet.column_dimensions[chr(i + ord(start))].width = max_len

    def fill_header(self, boxes, name_column, is_bold):
        """Заполняет строку с заголовками в excel

        Args:
            boxes (List): Список номеров столбцов в excel
            name_column (List):  Список столбцов с названиями
            is_bold (bool): Жирный шрифт или нет
        """

        i = 0
        for column in boxes:
            for box in column:
                box.value = name_column[i]
                box.font = Font(bold=is_bold)
                box.border = self.border
                i += 1

    def generate_excel(self, list_name, data, header, start, end, type=""):
        """Заполняет данными лист в excel

        Args:
            list_name (str): Название листа
            data (List): Данные со статистикой
            header (List): Список столбцов с названиями
            start (str): Номер первого столбца в excel
            end (str): Номер последнего столбца в excel
            type (str): Тип для блоков
        """

        self.ws = self.wb[list_name]
        self.fill_header(list(self.ws[start + '1': end + '1']), header, True)
        for i in range(len(data)):
            keys = list(data[i].keys())
            for j in range(len(keys)):
                self.ws[start + str(j + 2)].value = keys[j]
                self.ws[start + str(j + 2)].border = self.border
                self.ws[chr(i + ord(start) + 1) + str(j + 2)].value = data[i][keys[j]]
                self.ws[chr(i + ord(start) + 1) + str(j + 2)].border = self.border
                if type:
                    self.ws[chr(i + ord(start) + 1) + str(j + 2)].number_format = FORMAT_PERCENTAGE_00
        self.set_column_width(header, self.ws, len(list(data[0].keys())) + 1, start)

    def save_excel(self):
        """Сохраняет excel файл"""

        self.wb.save("report.xlsx")

    def generate_image(self, salary_by_year, salary_by_year_for_profession, count_by_year, count_by_year_for_profession, sum_salary_by_year_for_city, fraction_by_city):
        """Герерирует изображение со всеми графиками

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            salary_by_year_for_profession (dict): Словарь средней зарплаты по годам для профессии
            count_by_year (dict): Словарь количества вакансий по годам
            count_by_year_for_profession (dict): Cловарь количества ва кансий по годам для профессии
            sum_salary_by_year_for_city (dict): Уровень запрлат по городам
            fraction_by_city (dict): Доля вакансий по годам
        """

        width = 0.35
        x = np.array(list(salary_by_year.keys()))
        self.ax[0, 0].bar(x - width / 2, list(salary_by_year.values()), width, label='Средняя з/п')
        self.ax[0, 0].bar(x + width / 2, list(salary_by_year_for_profession.values()), width, label=f'з/п {profession}')
        self.ax[0, 0].set_title('Уровень зарплат по годам')
        self.ax[0, 0].set_xticks(x)
        self.ax[0, 0].set_xticklabels(list(salary_by_year.keys()), rotation=90, fontsize=8)
        self.ax[0, 0].legend(prop={'size': 8})
        self.ax[0, 0].grid(axis='y')

        self.ax[0, 1].bar(x - width / 2, list(count_by_year.values()), width, label='Количество вакансий')
        self.ax[0, 1].bar(x + width / 2, list(count_by_year_for_profession.values()), width, label=f'Количество вакансий {profession}')
        self.ax[0, 1].set_title('Количество вакансий по годам')
        self.ax[0, 1].set_xticks(x)
        self.ax[0, 1].set_xticklabels(list(salary_by_year.keys()), rotation=90, fontsize=8)
        self.ax[0, 1].legend(prop={'size': 8})
        self.ax[0, 1].grid(axis='y')

        label = []
        for item in list(sum_salary_by_year_for_city.keys()):
            if "-" in item:
                label.append(item.replace("-", "-\n"))
            elif " " in item:
                label.append(item.replace(" ", "\n"))
            else:
                label.append(item)

        y = np.array(label)
        self.ax[1, 0].barh(y, list(sum_salary_by_year_for_city.values()))
        self.ax[1, 0].set_yticks(y, labels=label, fontsize=6)
        self.ax[1, 0].invert_yaxis()
        self.ax[1, 0].set_title('Уровень зарплат по городам')
        self.ax[1, 0].grid(axis='x')

        summ = sum(list(fraction_by_city.values())[10:])

        fraction_by_city = dict(list(fraction_by_city.items())[0:10])
        fraction_by_city["Другие"] = summ

        self.ax[1, 1].pie(list(fraction_by_city.values()), labels=list(fraction_by_city.keys()), textprops={'fontsize': 6})
        self.ax[1, 1].axis("equal")
        self.ax[1, 1].set_title('Доля вакансий по городам')

        plt.tight_layout()
        plt.savefig('graph.png', dpi=100)

    def create_pdf(self):
        """Генерирует pdf файл используя html вёрстку"""

        options = {
            "enable-local-file-access": None
        }

        xfile = openpyxl.load_workbook("report.xlsx")
        data = xfile['Статистика по годам']
        data2 = xfile['Статистика по городам']

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(self.html.render({'profession': self.profession, 'first': data, 'second': data2}), 'report.pdf', configuration=config, options=options)



class DataSet:
    """Класс для создания статистики обработкой csv файла

    Attributes:
        file_name (str): Название csv файла
        filtering_parameter (str): Название профессии
    """

    def __init__(self, file_name, filtering_parameter):
        """Инициализирует объект DataSet

        Args:
            file_name (str): Название csv файла
            filtering_parameter (str): Название профессии
        """

        self.file_name = file_name
        self.filtering_parameter = filtering_parameter

    def csv_reader(self):
        """Считывает csv файл"""

        data = list(csv.reader(open(self.file_name, encoding="utf-8-sig")))
        if len(data) == 0:
            print("Пустой файл")
            quit()
        data_header = data[0]
        vacancies = [x for x in data[1:]]
        self.csv_splitter(vacancies)
        self.csv_filer(data_header, vacancies)

    def csv_splitter(self, vacancies):
        for i in range(2007, 2023):
            with open("csv_data/" + str(i) + ".csv", "a", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"])
        for item in vacancies:
            year = item[-1].split("-")[0]
            with open("csv_data/" + year + ".csv", "a", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(item)

    def splitter_csv_reader(self, file_name):
        data = list(csv.reader(open("csv_data/"+str(file_name)+".csv", encoding="utf-8-sig")))
        data_header = data[0]
        vacancies = [x for x in data[1:]]
        return self.csv_filer(data_header, vacancies)

    def csv_filer(self, list_naming, reader):
        """Фильтрация вакансий в csv файле

        Args:
            list_naming (List): Список шапки csv файла
            reader (List): Список вакансий
        """

        dict_count_by_year = {}
        dict_sum_salary_by_year = {}
        dict_count_by_year_for_profession = {}
        dict_sum_salary_by_year_for_profession = {}

        reader = [x for x in reader if not "" in x and len(x) == len(list_naming)]
        filter_list_vacancies = []
        for i in range(len(reader)):
            dict_vacancy = {}
            for j in range(len(list_naming)):
                vacancies_element = reader[i][j]
                if list_naming[j] == "key_skills":
                    vacancies_element = vacancies_element.split("\n")
                    vacancies_element = [x.strip() for x in vacancies_element]
                dict_vacancy[list_naming[j]] = vacancies_element
            dict_vacancy["salary_from"] = Salary([dict_vacancy["salary_from"], dict_vacancy["salary_to"], dict_vacancy["salary_currency"]])
            dict_vacancy = self.formatter(dict_vacancy)
            filter_list_vacancies.append(Vacancy(dict_vacancy))

        for item in filter_list_vacancies:
            self.set_statistics(item.vacansy_dict, dict_count_by_year, dict_sum_salary_by_year, dict_count_by_year_for_profession, dict_sum_salary_by_year_for_profession)

        if len(filter_list_vacancies) == 0:
            print("Нет данных")
            quit()

        for item in dict_sum_salary_by_year:
            dict_sum_salary_by_year[item] = dict_sum_salary_by_year[item] // dict_count_by_year[item]
            if item not in dict_sum_salary_by_year_for_profession:
                dict_sum_salary_by_year_for_profession[item] = 0
                dict_count_by_year_for_profession[item] = 0
                continue
            dict_sum_salary_by_year_for_profession[item] = dict_sum_salary_by_year_for_profession[item] // dict_count_by_year_for_profession[item]

        return [dict_sum_salary_by_year, dict_count_by_year, dict_sum_salary_by_year_for_profession, dict_count_by_year_for_profession]

    def set_statistics(self, list_vacancy, dict_count_by_year, dict_sum_salary_by_year, dict_count_by_year_for_profession, dict_sum_salary_by_year_for_profession):
        """Рассчитывает всю статистику

        Args:
            list_vacancy (Dict): Словарь с элементами одной вакансии
        """

        if list_vacancy["published_at"] not in dict_count_by_year:
            dict_count_by_year[list_vacancy["published_at"]] = 1
            dict_sum_salary_by_year[list_vacancy["published_at"]] = list_vacancy["salary_from"]
        else:
            dict_count_by_year[list_vacancy["published_at"]] += 1
            dict_sum_salary_by_year[list_vacancy["published_at"]] += list_vacancy["salary_from"]

        if list_vacancy["published_at"] not in dict_count_by_year_for_profession and self.filtering_parameter in list_vacancy["name"]:
            dict_count_by_year_for_profession[list_vacancy["published_at"]] = 1
            dict_sum_salary_by_year_for_profession[list_vacancy["published_at"]] = list_vacancy["salary_from"]
        elif self.filtering_parameter in list_vacancy["name"]:
            dict_count_by_year_for_profession[list_vacancy["published_at"]] += 1
            dict_sum_salary_by_year_for_profession[list_vacancy["published_at"]] += list_vacancy["salary_from"]

    def formatter(self, row):
        """Форматирует отдельные элементы вакансии для лучшего отображения

        Args:
            row (dict): Словарь с элементами одной вакансии
        :return:
        """

        date = row["published_at"][:10].split("-")
        row["published_at"] = int(date[0])
        row["salary_from"] = int((float(row["salary_from"].salary_from) + float(row["salary_from"].salary_to)) // 2 * currency_to_rub[row["salary_from"].salary_currency])
        return row

class Vacancy:
    """Класс со всеми атрибутами одной вакансии

    Attributes:
        vacansy_dict (dict): Словарь со всеми элементами одной вакансии
    """

    def __init__(self, list):
        """Инициализирует объект Vacansy

        Args:
            list (Dict): Словарь со всеми элементами одной вакансии
        """

        self.vacansy_dict = list

    @staticmethod
    def get_vacancy_type(list):
        return type(Vacancy(list)).__name__

    @staticmethod
    def get_vacancy_fields(list):
        objects = []
        for key in list:
            objects.append(list[key])
        return objects

    @staticmethod
    def get_vacancy_field(list, field_name):
        return Vacancy(list).vacansy_dict[field_name]

class Salary:
    """Класс со всеми атрибутами зарплаты

    Attributes:
        salary_from (str): Нижняя граница зарплаты
        salary_to (str): Верхняя граница зарплаты
        salary_currency (str): Валюта
    """

    def __init__(self, list):
        """Инициализирует объект Salary
            list (List): Лист со всеми значениями для атрибутов
        """

        self.salary_from, \
        self.salary_to,\
        self.salary_currency = list

    @staticmethod
    def get_salary_type(list):
        return type(Salary(list)).__name__

    @staticmethod
    def get_salary_from(list):
        return Salary(list).salary_from

    @staticmethod
    def get_salary_to(list):
        return Salary(list).salary_to

    @staticmethod
    def get_salary_currency(list):
        return Salary(list).salary_currency

    @staticmethod
    def get_salary(list):
        salary = Salary(list)
        return int((float(salary.salary_from) + float(salary.salary_to)) // 2 * currency_to_rub[salary.salary_currency])

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

dict_count_by_year = {}
dict_sum_salary_by_year = {}
dict_count_by_year_for_profession = {}
dict_sum_salary_by_year_for_profession = {}

if __name__ == '__main__':
    file_name = input("Введите название файла: ")
    profession = input("Введите название профессии: ")
    start = datetime.now()
    data_set = DataSet(file_name, profession)
    """pool = Pool(14)
    results = pool.map(data_set.splitter_csv_reader, range(2007, 2023))"""
    with concurrent.futures.ProcessPoolExecutor() as executor:
        for result in executor.map(data_set.splitter_csv_reader, range(2007, 2023)):
            dict_sum_salary_by_year |= result[0]
            dict_count_by_year |= result[1]
            dict_sum_salary_by_year_for_profession |= result[2]
            dict_count_by_year_for_profession |= result[3]
    print(f"Динамика уровня зарплат по годам: {dict_sum_salary_by_year}")
    print(f"Динамика количества вакансий по годам: {dict_count_by_year}")
    print(f"Динамика уровня зарплат по годам для выбранной профессии: {dict_sum_salary_by_year_for_profession}")
    print(f"Динамика количества вакансий по годам для выбранной профессии: {dict_count_by_year_for_profession}")
    print(datetime.now() - start)